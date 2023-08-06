import itertools
import pathlib
from abc import ABCMeta
from abc import abstractmethod
from decimal import Decimal
from typing import Iterable
from typing import Sequence
from typing import Tuple
from typing import Optional
from typing import Mapping

import ujson as json
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

from ._models import StudentScores
from ._typing import Modules
from ._modules import ModulesIndex


class ScoresReaderError(Exception):
    pass


class UnknownFileFormatError(ScoresReaderError):
    pass


class InvalidFileFormatError(ScoresReaderError):
    pass


class AbstractScoresReader(metaclass=ABCMeta):
    @abstractmethod
    def read(self, file: pathlib.Path, modules: Modules) -> Iterable[StudentScores]:
        raise NotImplementedError()


class JsonScoresReader(AbstractScoresReader):
    """
    WARNING: реализация неэффективная, так как загружает json-файл весь
     в память
    """

    def read(self, file: pathlib.Path, modules: Modules) -> Iterable[StudentScores]:
        modules_index = ModulesIndex(modules)
        course_structure = json.load(file.open('rb'))

        try:
            units = course_structure['course']['units']
            students = course_structure['students']
        except KeyError:
            raise InvalidFileFormatError()

        requested_modules_ids = set()
        for unit in units:
            if unit['title'] in modules_index:
                for slide in unit['slides']:
                    requested_modules_ids.add(slide['id'])

        for student in students:
            scores = Decimal(0)

            for slide in student['slides_scores']:
                if slide['slide_id'] in requested_modules_ids:
                    scores += Decimal(slide['score'])

            yield StudentScores(
                student=student['name'],
                scores=scores,
            )


class XlsxScoresReader(AbstractScoresReader):
    def read(self, file: pathlib.Path, modules: Modules) -> Iterable[StudentScores]:
        workbook = load_workbook(str(file), read_only=True)

        try:
            worksheet = workbook.active
            required_columns = self._get_required_column_numbers(
                header_row=next(worksheet.iter_rows()),
                max_column=worksheet.max_column,
                modules=modules,
            )
            for row in worksheet.iter_rows(min_row=4):
                yield StudentScores(
                    student=row[0].value,
                    scores=self._scores_summary(row, required_columns),
                )
        finally:
            workbook.close()

    @classmethod
    def _get_required_column_numbers(
        cls,
        header_row: Tuple[Cell, ...],
        max_column: int,
        modules: Modules,
    ) -> Tuple[int, ...]:
        modules_index = ModulesIndex(modules)
        required_columns = []
        cell_modules = cls._get_all_modules(header_row)
        cells_iter = itertools.islice(cell_modules, max_column)

        for cell in cells_iter:
            # Костыльное решение с проверкой None
            if cell.value in modules_index:
                required_columns.append(cell.column)

        return tuple(required_columns)

    @staticmethod
    def _get_all_modules(header_row: Tuple[Cell, ...]) -> Tuple[Cell, ...]:
        for cell in itertools.islice(header_row, 4, None, None):
            if cell.value is not None and cell.value != '':
                yield cell

    @staticmethod
    def _scores_summary(
        row: Tuple[Cell, ...],
        required_columns: Sequence[int],
    ) -> Decimal:
        summary = Decimal(0)

        for required_column in required_columns:
            summary += Decimal(row[required_column - 1].value)

        return summary


class AutoScoresReader(AbstractScoresReader):
    default_readers_by_extensions = {
        'xlsx': XlsxScoresReader(),
        'json': JsonScoresReader(),
    }

    def __init__(
        self,
        readers_by_extension: Optional[Mapping[str, AbstractScoresReader]] = None
    ):
        self._readers_by_extension = readers_by_extension or self.default_readers_by_extensions

    def read(self, file: pathlib.Path, modules: Modules) -> Iterable[StudentScores]:
        parts = file.name.rsplit('.', maxsplit=1)

        if len(parts) < 2:
            raise UnknownFileFormatError()

        _, extension = parts

        try:
            reader = self._readers_by_extension[extension]
        except KeyError:
            raise UnknownFileFormatError(extension)

        return reader.read(file, modules)
