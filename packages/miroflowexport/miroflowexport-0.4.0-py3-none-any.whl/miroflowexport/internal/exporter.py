import csv
from email import header
import openpyxl as xl
from openpyxl.utils import get_column_letter
from openpyxl.formatting import Rule
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.differential import DifferentialStyle

CSV_DELIMITER = ','

BG_COLOR_HORIZON = 'E2E5DE'
BG_COLOR_PLAN = '1AA7EC'

COLOR_PROGRESS_0_NOT_STARTED = '1AA7EC'
COLOR_PROGRESS_1_STARTED = 'FCF4A3'
COLOR_PROGRESS_2_WORKING = 'F8E473'
COLOR_PROGRESS_3_NEARLY_DONE = '80b280'
COLOR_PROGRESS_4_DONE = '33b333'

TASK_FILL_HORIZON = 'o'
TASK_FILL_PLAN = 'x'

def apply_conditional_progress_coloring(ws, base_formula_plan, formula_range, color, reference_column, upper_threshold):
    color_fill_plan = PatternFill(bgColor = color)
    style_plan = DifferentialStyle(fill = color_fill_plan, font = Font(color = color))
    rule_plan = Rule(type="expression", dxf=style_plan, stopIfTrue=True)
    rule_plan.formula = ["AND({}, {})".format(base_formula_plan, "{}<={}".format(reference_column, upper_threshold))]
    ws.conditional_formatting.add(formula_range, rule_plan)

def tasks_to_csv(log, tasks, path = "tmp/list.csv"):
    log.debug("Writing cards to this file: {}".format(path))
    try:
        with open(path, newline = '', mode = 'w', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile, delimiter = CSV_DELIMITER)
            writer.writerow([
                "ID", 
                "Aufgabe", 
                "Aufwand", 
                "ES", 
                "LF",
            ])
            for task in tasks.values():
                writer.writerow([
                    task.id(),
                    task.name(),
                    task.effort(),
                    task.earliest_start(),
                    task.latest_finish(),
                ])
        log.info("Entries written to {}.".format(path))
    except:
        log.exception("Writing entries to {} failed.".format(path))

def dependencies_to_csv(log, tasks, path = "tmp/dependencies.csv"):
    log.debug("Writing dependencies to this file: {}".format(path))
    try:
        with open(path, newline = '', mode = 'w', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile, delimiter = CSV_DELIMITER)
            writer.writerow(["ID_A", "Predecessor", "ID_B", "Successor"])
            for task in tasks.values():
                for next in task.successors():
                    writer.writerow([
                        task.id(),
                        task.name(),
                        next.id(),
                        next.name()
                    ])
        log.info("Entries written to {}.".format(path))
    except:
        log.exception("Writing entries to {} failed.".format(path))

def tasks_to_excel(log, tasks_dict, path = "tmp/export.xlsx", sheet_tasks = "tasks"):
    log.debug("Creating Excel export for {} tasks ...".format(len(tasks_dict.values())))
    wb = xl.Workbook()

    ws = wb.active
    ws.title = sheet_tasks

    week_start = min([
        task.earliest_start()
        for task in tasks_dict.values()
    ])
    week_end = max([
        task.latest_finish()
        for task in tasks_dict.values()
    ])
    cols_weeks = [
        "Week {}".format(week)
        for week in range(week_start, week_end + 1)
    ]

    header_row = [
        "Task ID",
        "Task Name",
        "ES",
        "LF",
        "Effort (weeks)",
        "Horizon Start",
        "Horizon End",
        "Task Start",
        "Task Finish",
        "Progress",
    ]
    col_num_progress = header_row.index("Progress") + 1
    col_letter_progress = get_column_letter(col_num_progress)

    groups_to_hide = [
        # Hide Task ID
        (get_column_letter(header_row.index("Task ID") + 1), get_column_letter(header_row.index("Task ID") + 1)),
        # Hide suporting information
        (get_column_letter(header_row.index("ES") + 1), get_column_letter(header_row.index("Task Finish") + 1)),
    ]

    num_cols_before_weeks = len(header_row)
    header_row.extend(cols_weeks)

    ws.append(header_row)

    col_letters_for_week_cols = [
        get_column_letter(week_col)
        for week_col in range(num_cols_before_weeks + 1, num_cols_before_weeks + 1 + len(cols_weeks))
    ]

    current_row = 1
    for task_id in tasks_dict:
        current_row += 1
        task = tasks_dict[task_id]
        content_row = [
            task.id(),
            task.name(),
            task.earliest_start(),
            task.latest_finish(),
            task.effort(),
            week_start,
            week_end,
            task.start(),
            task.end(),
            task.progress(),
        ]
        content_row.extend([
            "" if week < task.earliest_start() or week > task.latest_finish() else TASK_FILL_PLAN if week >= task.start() and week <= task.end() else TASK_FILL_HORIZON
            for week in range(week_start, week_end + 1)
        ])
        ws.append(content_row)
        ws["{}{}".format(col_letter_progress, current_row)].number_format = "0%"

    formula_base = "{}2=".format(col_letters_for_week_cols[0])
    log.debug("Formula base for conditional formatting: {}".format(formula_base))

    formula_range = "{}2:{}{}".format(
        col_letters_for_week_cols[0],
        col_letters_for_week_cols[-1],
        len(tasks_dict.keys()) + 1
    )
    log.debug("Formula range for conditional formatting: {}".format(formula_range))

    formula_horizon = '{}"{}"'.format(formula_base, TASK_FILL_HORIZON)
    formula_plan = '{}"{}"'.format(formula_base, TASK_FILL_PLAN)
    log.debug("Formula for horizon formatting: {}".format(formula_horizon))
    log.debug("Formula for plan formatting:    {}".format(formula_plan))

    font_horizon = Font(color = BG_COLOR_HORIZON)
    
    color_fill_horizon = PatternFill(bgColor = BG_COLOR_HORIZON)
    style_horizon = DifferentialStyle(fill = color_fill_horizon, font = font_horizon)
    rule_horizon = Rule(type="expression", dxf=style_horizon, stopIfTrue=True)
    rule_horizon.formula = [formula_horizon]
    ws.conditional_formatting.add(formula_range, rule_horizon)

    ref_column = "${}2".format(col_letter_progress)
    for color, threshold in [
        (COLOR_PROGRESS_0_NOT_STARTED, 0),
        (COLOR_PROGRESS_1_STARTED, 0.25),
        (COLOR_PROGRESS_2_WORKING, 0.75),
        (COLOR_PROGRESS_3_NEARLY_DONE, 0.99),
        (COLOR_PROGRESS_4_DONE, 1),
    ]:
        apply_conditional_progress_coloring(
            ws = ws,
            base_formula_plan = formula_plan,
            formula_range = formula_range,
            color = color,
            reference_column = ref_column,
            upper_threshold = threshold,
        )

    for col_letter in col_letters_for_week_cols:
        ws.column_dimensions[col_letter].width = 3

    for (col_start, col_end) in groups_to_hide:
        ws.column_dimensions.group(col_start, col_end, hidden = True)

    log.debug("Saving Excel workbook to {} ...".format(path))
    try:
        wb.save(path)
    except:
        log.exception("Saving Excel workbook failed. See exception for details.")
        return False

    return True

if __name__ == "__main__":
    wb = xl.load_workbook("tmp/export.xlsx")
    ws = wb.active

    col_progress = ws.column_dimensions['J']
    print("Number format: '{}'".format(col_progress.number_format))