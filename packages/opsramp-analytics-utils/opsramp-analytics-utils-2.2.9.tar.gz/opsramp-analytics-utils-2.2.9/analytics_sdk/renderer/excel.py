# import random
# import dateutil.parser
# from datetime import datetime
# import logging
#
# from openpyxl import Workbook
# from openpyxl.styles import PatternFill, Font, Alignment
# from openpyxl.worksheet.table import Table, TableStyleInfo
# from openpyxl.utils import get_column_letter
# from openpyxl.chart import (
#     DoughnutChart,
#     ScatterChart,
#     BarChart,
#     Reference,
#     Series,
# )
# from openpyxl.chart.series import DataPoint
#
#
# class ExcelRenderer:
#     ROW_START = 3
#     logging.error(f'{ROW_START} <******  enter into ExcelRenderer /render/excel.py  ******>')
#     def __init__(self, analysis_run):
#         """
#         :param analysis_run: AnalysisRun
#         """
#         self.wb = Workbook()
#         self.analysis_run = analysis_run
#
#     def _load_data(self, ws, data, row_start, col_start, adjust_column=True, x_axis_date_format=None):
#         column_widths = {}
#         for row_delta, row in enumerate(data):
#             for col_delta, val in enumerate(row):
#                 col = col_start + col_delta
#                 cell = ws.cell(row=row_start+row_delta, column=col)
#                 if x_axis_date_format and row_delta > 0 and col_delta == 0:
#                     cell.value = datetime.strptime(val, x_axis_date_format).date()
#                 else:
#                     cell.value = val
#
#                 if col in column_widths:
#                     column_widths[col] = max(len(str(val)), column_widths[col])
#                 else:
#                     column_widths[col] = max(len(str(val)), ws.column_dimensions[get_column_letter(col)].width)
#
#         # handle cell width
#         if adjust_column:
#             for col, column_width in column_widths.items():
#                 ws.column_dimensions[get_column_letter(col)].width = column_width + 2
#
#     def add_table(self, ws, table_data):
#         """
#         add table component
#         """
#         row_start = ExcelRenderer.ROW_START + table_data['start-row'] + 1
#         col_start = table_data['start-col']
#         self._load_data(ws, table_data['data'], row_start, col_start)
#         row_span = len(table_data['data'])
#         col_span = len(table_data['data'][0])
#
#         ref = f'{get_column_letter(col_start)}{row_start}:{get_column_letter(col_start+col_span-1)}{row_start+row_span-1}'
#         table_name = f'Table{random.randint(0, 100)}'
#
#         table = Table(displayName=table_name, ref=ref, tableStyleInfo=TableStyleInfo(name="TableStyleMedium9"))
#         ws.add_table(table)
#
#     def add_doughnut_chart(self, ws, chart_data):
#         """
#         add doughnut chart component
#         """
#         row_start = ExcelRenderer.ROW_START + chart_data['start-row']
#         col_start = chart_data['start-col']
#         self._load_data(ws, chart_data['data'], row_start, col_start, False)
#         row_span = len(chart_data['data'])
#
#         chart = DoughnutChart(holeSize=50)
#         labels = Reference(ws, min_col=col_start, min_row=row_start+1, max_row=row_start+row_span-1)
#         data = Reference(ws, min_col=col_start+1, min_row=row_start, max_row=row_start+row_span-1)
#         chart.add_data(data, titles_from_data=True)
#         chart.set_categories(labels)
#         chart.title = chart_data['chart-title']
#         chart.style = 12
#         chart.width = chart_data.get('width', 11)
#         chart.height = chart_data.get('height', 5)
#
#         slices = [DataPoint(idx=i) for i in range(row_span-1)]
#         chart.series[0].data_points = slices
#         colors = ["0077C8", "32a3df", "673ab7", "9c27b0"]
#
#         for idx, slice in enumerate(slices):
#             slice.graphicalProperties.solidFill = colors[idx % 4]
#
#         ws.add_chart(chart, chart_data['chart-position'])
#
#     def add_bar_chart(self, ws, chart_data):
#         """
#         add bar chart component
#         """
#         row_start = ExcelRenderer.ROW_START + chart_data['start-row']
#         col_start = chart_data['start-col']
#         self._load_data(ws, chart_data['data'], row_start, col_start, False)
#         row_span = len(chart_data['data'])
#
#         chart = BarChart()
#         chart.type = "col"
#         chart.style = 10
#         chart.title = chart_data['chart-title']
#         chart.x_axis.title = chart_data.get('x-axis-title')
#         chart.y_axis.title = chart_data.get('y-axis-title')
#
#         labels = Reference(ws, min_col=col_start, min_row=row_start+1, max_row=row_start+row_span-1)
#         data = Reference(ws, min_col=col_start+1, min_row=row_start, max_row=row_start+row_span-1)
#         chart.add_data(data, titles_from_data=True)
#         chart.set_categories(labels)
#         chart.style = 4
#         chart.legend = None
#         chart.width = chart_data.get('width', 17)
#         chart.height = chart_data.get('height', 8.5)
#
#         s1 = chart.series[0]
#         s1.graphicalProperties.line.solidFill = "0077C8"
#         s1.graphicalProperties.solidFill = "0077C8"
#
#         ws.add_chart(chart, chart_data['chart-position'])
#
#     def add_scatter_chart(self, ws, chart_data):
#         """
#         add scatter chart component
#         """
#         row_start = ExcelRenderer.ROW_START + chart_data['start-row']
#         col_start = chart_data['start-col']
#         x_axis_date_format = chart_data.get('x-axis-date-format')
#         self._load_data(ws, chart_data['data'], row_start, col_start, False, x_axis_date_format)
#         row_span = len(chart_data['data'])
#
#         chart = ScatterChart()
#         chart.style = 5
#         chart.legend = None
#         chart.title = chart_data['chart-title']
#         chart.x_axis.title = chart_data.get('x-axis-title')
#         chart.y_axis.title = chart_data.get('y-axis-title')
#         chart.width = chart_data.get('width', 15)
#         chart.height = chart_data.get('height', 7.5)
#
#         labels = Reference(ws, min_col=col_start, min_row=row_start+1, max_row=row_start+row_span-1)
#         data = Reference(ws, min_col=col_start+1, min_row=row_start, max_row=row_start+row_span-1)
#         series = Series(data, labels, title_from_data=True)
#         chart.series.append(series)
#
#         s1 = chart.series[0]
#         s1.marker.symbol = "circle"
#         s1.marker.graphicalProperties.solidFill = "0077C8"  # Marker filling
#         s1.marker.graphicalProperties.line.solidFill = "0077C8"  # Marker outline
#         s1.graphicalProperties.line.solidFill = "0077C8"
#         s1.graphicalProperties.line.width = 24050  # width in EMUs
#         if x_axis_date_format:
#             chart.x_axis.number_format = 'm/d'
#             chart.x_axis.majorTimeUnit = "days"
#
#         ws.add_chart(chart, chart_data['chart-position'])
#
#     def generate_summary(self):
#         """
#         add summary sheet
#         """
#         ws = self.wb.active
#         ws.title = 'Summary'
#
#         for row_idx in range(1, 30):
#             row = ws.row_dimensions[row_idx]
#             row.fill = PatternFill("solid", fgColor="eeeeee")
#
#         ws.column_dimensions['B'].width = 24
#         ws.column_dimensions['C'].width = 24
#
#         ws.merge_cells('B3:Z4')
#         cell = ws['B3']
#         cell.value = self.analysis_run.analysis.app.name
#         cell.font = Font(color="00598B", bold=True)
#         cell.alignment = Alignment(vertical="center")
#         cell.fill = PatternFill("solid", fgColor="eeeeee")
#
#         cell = ws['B5']
#         cell.value = 'Analysis Parameters'
#         cell.font = Font(color="00598B")
#         cell.fill = PatternFill("solid", fgColor="eeeeee")
#
#         cell = ws['B6']
#         cell.value = 'Client'
#         cell.alignment = Alignment(horizontal="right")
#         cell.font = Font(color="00598B")
#         cell.fill = PatternFill("solid", fgColor="eeeeee")
#
#         cell = ws['C6']
#         cell.value = self.analysis_run.analysis.name
#         cell.font = Font(color="000000")
#         cell.fill = PatternFill("solid", fgColor="eeeeee")
#
#         cell = ws['B7']
#         cell.value = 'Analysis Period'
#         cell.alignment = Alignment(horizontal="right")
#         cell.font = Font(color="00598B")
#         cell.fill = PatternFill("solid", fgColor="eeeeee")
#
#         cell = ws['C7']
#         params = self.analysis_run.params
#         start_date = dateutil.parser.isoparse(params["start_date"])
#         end_date = dateutil.parser.isoparse(params["end_date"])
#         cell.value = f'{start_date.strftime("%B %d, %Y")} - {end_date.strftime("%B %d, %Y")}'
#         cell.font = Font(color="000000")
#         cell.fill = PatternFill("solid", fgColor="eeeeee")
#
#         cell = ws['B8']
#         cell.value = 'Analysis'
#         cell.font = Font(color="00598B")
#         cell.fill = PatternFill("solid", fgColor="eeeeee")
#
#         cell = ws['B9']
#         cell.value = 'Analysis Name'
#         cell.alignment = Alignment(horizontal="right")
#         cell.font = Font(color="00598B")
#         cell.fill = PatternFill("solid", fgColor="eeeeee")
#
#         cell = ws['C9']
#         cell.value = self.analysis_run.analysis.name
#         cell.font = Font(color="000000")
#         cell.fill = PatternFill("solid", fgColor="eeeeee")
#
#         cell = ws['B10']
#         cell.value = 'Run date'
#         cell.alignment = Alignment(horizontal="right")
#         cell.font = Font(color="00598B")
#         cell.fill = PatternFill("solid", fgColor="eeeeee")
#
#         cell = ws['C10']
#         cell.value = self.analysis_run.date_launched.strftime('%H:%M %p %Z, %B %d, %Y')
#         cell.font = Font(color="000000")
#         cell.fill = PatternFill("solid", fgColor="eeeeee")
#
#         cell = ws['B11']
#         cell.value = 'Complete date'
#         cell.alignment = Alignment(horizontal="right")
#         cell.font = Font(color="00598B")
#         cell.fill = PatternFill("solid", fgColor="eeeeee")
#
#         cell = ws['C11']
#         cell.value = self.analysis_run.date_completed.strftime('%H:%M %p %Z, %B %d, %Y')
#         cell.font = Font(color="000000")
#         cell.fill = PatternFill("solid", fgColor="eeeeee")
#
#         cell = ws['B12']
#         cell.value = 'User'
#         cell.alignment = Alignment(horizontal="right")
#         cell.font = Font(color="00598B")
#         cell.fill = PatternFill("solid", fgColor="eeeeee")
#
#         # cell = ws['C12']
#         # user = self.analysis_run.analysis.app.installed_by
#         # cell.value = f'{user.first_name} {user.last_name}'
#         # cell.font = Font(color="000000")
#         # cell.fill = PatternFill("solid", fgColor="eeeeee")
#
#         cell = ws['B13']
#         cell.value = 'App version'
#         cell.alignment = Alignment(horizontal="right")
#         cell.font = Font(color="00598B")
#         cell.fill = PatternFill("solid", fgColor="eeeeee")
#
#         cell = ws['C13']
#         cell.value = self.analysis_run.analysis.app.version
#         cell.font = Font(color="000000")
#         cell.fill = PatternFill("solid", fgColor="eeeeee")
#
#     def add_header(self, ws, sheet_data):
#         """
#         add header to the sheet
#         """
#         for row_idx in range(1, ExcelRenderer.ROW_START):
#             row = ws.row_dimensions[row_idx]
#             row.fill = PatternFill("solid", fgColor="eeeeee")
#
#         ws.merge_cells('B1:Z2')
#         cell = ws['B1']
#         cell.value = self.analysis_run.analysis.app.name
#         cell.font = Font(color="00598B", bold=True)
#         cell.fill = PatternFill("solid", fgColor="eeeeee")
#         cell.alignment = Alignment(vertical="center")
#
#     def add_component_title(self, ws, component_data):
#         """
#         add component title
#         """
#         cell = ws.cell(row=ExcelRenderer.ROW_START+component_data['start-row'], column=component_data['start-col'])
#         cell.value = component_data['title']
#         color = component_data.get('color', '00598B')
#         cell.font = Font(color=color, bold=True)
#
#     def add_component(self, ws, component_data):
#         if component_data['type'] == 'table':
#             if component_data.get('title'):
#                 self.add_component_title(ws, component_data)
#             self.add_table(ws, component_data)
#         elif component_data['type'] == 'doughnut-chart':
#             self.add_doughnut_chart(ws, component_data)
#         elif component_data['type'] == 'bar-chart':
#             self.add_bar_chart(ws, component_data)
#         elif component_data['type'] == 'scatter-chart':
#             self.add_scatter_chart(ws, component_data)
#
#     def render(self):
#         """
#         :return: Workbook
#         """
#         self.generate_summary()
#         run_data = self.analysis_run.get_result().get('excel-data', {})
#
#         for sheet in run_data['sheets']:
#             ws = self.wb.create_sheet(title=sheet['title'])
#             self.add_header(ws, sheet)
#
#             for component_data in sheet['components']:
#                 self.add_component(ws, component_data)
#
#         return self.wb


#==========================    Modified Code    ======================================
import random
import dateutil.parser
from datetime import datetime
import logging

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.chart import (
    DoughnutChart,
    ScatterChart,
    BarChart,
    Reference,
    Series,
)
from openpyxl.chart.series import DataPoint


class ExcelRenderer:
    ROW_START = 3
    logging.error(f'{ROW_START} <******  enter into ExcelRenderer /render/excel.py  ******>')
    def __init__(self, analysis_run):
        """
        :param analysis_run: AnalysisRun
        """
        self.wb = Workbook()
        self.analysis_run = analysis_run

    def _load_data(self, ws, data, row_start, col_start, adjust_column=True, x_axis_date_format=None):
        column_widths = {}
        for row_delta, row in enumerate(data):
            for col_delta, val in enumerate(row):
                col = col_start + col_delta
                cell = ws.cell(row=row_start+row_delta, column=col)
                if x_axis_date_format and row_delta > 0 and col_delta == 0:
                    cell.value = datetime.strptime(val, x_axis_date_format).date()
                else:
                    cell.value = val

                if col in column_widths:
                    column_widths[col] = max(len(str(val)), column_widths[col])
                else:
                    column_widths[col] = max(len(str(val)), ws.column_dimensions[get_column_letter(col)].width)

        # handle cell width
        if adjust_column:
            for col, column_width in column_widths.items():
                ws.column_dimensions[get_column_letter(col)].width = column_width + 2

    def add_table(self, ws, table_data):
        """
        add table component
        """
        row_start = ExcelRenderer.ROW_START + table_data['start-row'] + 1
        col_start = table_data['start-col']
        self._load_data(ws, table_data['data'], row_start, col_start)
        row_span = len(table_data['data'])
        col_span = len(table_data['data'][0])

        ref = f'{get_column_letter(col_start)}{row_start}:{get_column_letter(col_start+col_span-1)}{row_start+row_span-1}'
        table_name = f'Table{random.randint(0, 100)}'

        table = Table(displayName=table_name, ref=ref, tableStyleInfo=TableStyleInfo(name="TableStyleMedium9"))
        ws.add_table(table)

    def add_doughnut_chart(self, ws, chart_data):
        """
        add doughnut chart component
        """
        row_start = ExcelRenderer.ROW_START + chart_data['start-row']
        col_start = chart_data['start-col']
        self._load_data(ws, chart_data['data'], row_start, col_start, False)
        row_span = len(chart_data['data'])

        chart = DoughnutChart(holeSize=50)
        labels = Reference(ws, min_col=col_start, min_row=row_start+1, max_row=row_start+row_span-1)
        data = Reference(ws, min_col=col_start+1, min_row=row_start, max_row=row_start+row_span-1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        chart.title = chart_data['chart-title']
        chart.style = 12
        chart.width = chart_data.get('width', 11)
        chart.height = chart_data.get('height', 5)

        slices = [DataPoint(idx=i) for i in range(row_span-1)]
        chart.series[0].data_points = slices
        colors = ["0077C8", "32a3df", "673ab7", "9c27b0"]

        for idx, slice in enumerate(slices):
            slice.graphicalProperties.solidFill = colors[idx % 4]

        ws.add_chart(chart, chart_data['chart-position'])

    def add_bar_chart(self, ws, chart_data):
        """
        add bar chart component
        """
        row_start = ExcelRenderer.ROW_START + chart_data['start-row']
        col_start = chart_data['start-col']
        self._load_data(ws, chart_data['data'], row_start, col_start, False)
        row_span = len(chart_data['data'])

        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = chart_data['chart-title']
        chart.x_axis.title = chart_data.get('x-axis-title')
        chart.y_axis.title = chart_data.get('y-axis-title')

        labels = Reference(ws, min_col=col_start, min_row=row_start+1, max_row=row_start+row_span-1)
        data = Reference(ws, min_col=col_start+1, min_row=row_start, max_row=row_start+row_span-1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        chart.style = 4
        chart.legend = None
        chart.width = chart_data.get('width', 17)
        chart.height = chart_data.get('height', 8.5)

        s1 = chart.series[0]
        s1.graphicalProperties.line.solidFill = "0077C8"
        s1.graphicalProperties.solidFill = "0077C8"

        ws.add_chart(chart, chart_data['chart-position'])

    def add_scatter_chart(self, ws, chart_data):
        """
        add scatter chart component
        """
        row_start = ExcelRenderer.ROW_START + chart_data['start-row']
        col_start = chart_data['start-col']
        x_axis_date_format = chart_data.get('x-axis-date-format')
        self._load_data(ws, chart_data['data'], row_start, col_start, False, x_axis_date_format)
        row_span = len(chart_data['data'])

        chart = ScatterChart()
        chart.style = 5
        chart.legend = None
        chart.title = chart_data['chart-title']
        chart.x_axis.title = chart_data.get('x-axis-title')
        chart.y_axis.title = chart_data.get('y-axis-title')
        chart.width = chart_data.get('width', 15)
        chart.height = chart_data.get('height', 7.5)

        labels = Reference(ws, min_col=col_start, min_row=row_start+1, max_row=row_start+row_span-1)
        data = Reference(ws, min_col=col_start+1, min_row=row_start, max_row=row_start+row_span-1)
        series = Series(data, labels, title_from_data=True)
        chart.series.append(series)

        s1 = chart.series[0]
        s1.marker.symbol = "circle"
        s1.marker.graphicalProperties.solidFill = "0077C8"  # Marker filling
        s1.marker.graphicalProperties.line.solidFill = "0077C8"  # Marker outline
        s1.graphicalProperties.line.solidFill = "0077C8"
        s1.graphicalProperties.line.width = 24050  # width in EMUs
        if x_axis_date_format:
            chart.x_axis.number_format = 'm/d'
            chart.x_axis.majorTimeUnit = "days"

        ws.add_chart(chart, chart_data['chart-position'])

    def generate_summary(self):
        """
        add summary sheet
        """
        ws = self.wb.active
        ws.title = 'Summary'

        for row_idx in range(1, 30):
            row = ws.row_dimensions[row_idx]
            row.fill = PatternFill("solid", fgColor="eeeeee")

        ws.column_dimensions['B'].width = 24
        ws.column_dimensions['C'].width = 24

        ws.merge_cells('B3:Z4')
        cell = ws['B3']
        cell.value = "asset-insights"
        cell.font = Font(color="00598B", bold=True)
        cell.alignment = Alignment(vertical="center")
        cell.fill = PatternFill("solid", fgColor="eeeeee")

        cell = ws['B5']
        cell.value = 'Analysis Parameters'
        cell.font = Font(color="00598B")
        cell.fill = PatternFill("solid", fgColor="eeeeee")

        cell = ws['B6']
        cell.value = 'Client'
        cell.alignment = Alignment(horizontal="right")
        cell.font = Font(color="00598B")
        cell.fill = PatternFill("solid", fgColor="eeeeee")

        cell = ws['C6']
        cell.value = "Dec-11-analysis"
        cell.font = Font(color="000000")
        cell.fill = PatternFill("solid", fgColor="eeeeee")

        cell = ws['B7']
        cell.value = 'Analysis Period'
        cell.alignment = Alignment(horizontal="right")
        cell.font = Font(color="00598B")
        cell.fill = PatternFill("solid", fgColor="eeeeee")

        cell = ws['C7']
        # params = self.analysis_run.params
        start_date = "2021-12-10T12:00:00"
        end_date = "2021-12-11T12:00:00"
        cell.value = f'{start_date} - {end_date}'
        cell.font = Font(color="000000")
        cell.fill = PatternFill("solid", fgColor="eeeeee")

        cell = ws['B8']
        cell.value = 'Analysis'
        cell.font = Font(color="00598B")
        cell.fill = PatternFill("solid", fgColor="eeeeee")

        cell = ws['B9']
        cell.value = 'Analysis Name'
        cell.alignment = Alignment(horizontal="right")
        cell.font = Font(color="00598B")
        cell.fill = PatternFill("solid", fgColor="eeeeee")

        cell = ws['C9']
        cell.value = "Dec analysis"
        cell.font = Font(color="000000")
        cell.fill = PatternFill("solid", fgColor="eeeeee")

        cell = ws['B10']
        cell.value = 'Run date'
        cell.alignment = Alignment(horizontal="right")
        cell.font = Font(color="00598B")
        cell.fill = PatternFill("solid", fgColor="eeeeee")

        cell = ws['C10']
        cell.value = "2021-12-11T12:00:00"
        cell.font = Font(color="000000")
        cell.fill = PatternFill("solid", fgColor="eeeeee")

        cell = ws['B11']
        cell.value = 'Complete date'
        cell.alignment = Alignment(horizontal="right")
        cell.font = Font(color="00598B")
        cell.fill = PatternFill("solid", fgColor="eeeeee")

        cell = ws['C11']
        cell.value = "2021-12-12T12:00:00"
        cell.font = Font(color="000000")
        cell.fill = PatternFill("solid", fgColor="eeeeee")

        cell = ws['B12']
        cell.value = 'User'
        cell.alignment = Alignment(horizontal="right")
        cell.font = Font(color="00598B")
        cell.fill = PatternFill("solid", fgColor="eeeeee")

        # cell = ws['C12']
        # user = self.analysis_run.analysis.app.installed_by
        # cell.value = f'{user.first_name} {user.last_name}'
        # cell.font = Font(color="000000")
        # cell.fill = PatternFill("solid", fgColor="eeeeee")

        cell = ws['B13']
        cell.value = 'App version'
        cell.alignment = Alignment(horizontal="right")
        cell.font = Font(color="00598B")
        cell.fill = PatternFill("solid", fgColor="eeeeee")

        cell = ws['C13']
        cell.value = "11.O"
        cell.font = Font(color="000000")
        cell.fill = PatternFill("solid", fgColor="eeeeee")

    def add_header(self, ws, sheet_data):
        """
        add header to the sheet
        """
        for row_idx in range(1, ExcelRenderer.ROW_START):
            row = ws.row_dimensions[row_idx]
            row.fill = PatternFill("solid", fgColor="eeeeee")

        ws.merge_cells('B1:Z2')
        cell = ws['B1']
        cell.value = "Asset-Insights"
        cell.font = Font(color="00598B", bold=True)
        cell.fill = PatternFill("solid", fgColor="eeeeee")
        cell.alignment = Alignment(vertical="center")

    def add_component_title(self, ws, component_data):
        """
        add component title
        """
        cell = ws.cell(row=ExcelRenderer.ROW_START+component_data['start-row'], column=component_data['start-col'])
        cell.value = component_data['title']
        color = component_data.get('color', '00598B')
        cell.font = Font(color=color, bold=True)

    def add_component(self, ws, component_data):
        if component_data['type'] == 'table':
            if component_data.get('title'):
                self.add_component_title(ws, component_data)
            self.add_table(ws, component_data)
        elif component_data['type'] == 'doughnut-chart':
            self.add_doughnut_chart(ws, component_data)
        elif component_data['type'] == 'bar-chart':
            self.add_bar_chart(ws, component_data)
        elif component_data['type'] == 'scatter-chart':
            self.add_scatter_chart(ws, component_data)

    def render(self):
        """
        :return: Workbook
        """
        self.generate_summary()
        # run_data = self.analysis_run.get_result().get('excel-data', {})
        run_data = self.analysis_run
        for sheet in run_data['sheets']:
            ws = self.wb.create_sheet(title=sheet['title'])
            self.add_header(ws, sheet)

            for component_data in sheet['components']:
                self.add_component(ws, component_data)

        return self.wb




