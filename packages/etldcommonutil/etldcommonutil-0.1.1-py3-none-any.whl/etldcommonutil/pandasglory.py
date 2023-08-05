import csv
import pandas
from openpyxl import load_workbook

class PandasUtil:
    def __init__(self):
        pass
    # Converts query result to xlsx file
    def toxlsx(self, rows, header, filename_with_path):
            # New empty list called 'result'. This will be written to a file.
            result = list()
            # The row name is the first entry for each entity in the description tuple.
            column_names = list()
            for col in header:
                column_names.append(col)

            for row in rows:
                result.append(row)

            # Write result to file.
            df = pandas.DataFrame(result, columns=column_names)
            df.to_excel(filename_with_path, sheet_name='Dupe Records', index=False)

    # Converts query result to csv file
    def tocsv(self, rows, header, filename_with_path):
        # New empty list called 'result'. This will be written to a file.
        result = list()
        # The row name is the first entry for each entity in the description tuple.
        column_names = list()
        for col in header:
            column_names.append(col)
        result.append(column_names)

        for rrr in rows:
            result.append(rrr)

        # Write result to file.
        with open(filename_with_path, 'w', newline='') as csvfile:
            csvwriter = csv.writer(csvfile, delimiter=',', quotechar='"', quoting=csv.QUOTE_ALL)
            for row in result:
                csvwriter.writerow(row)

    # Converts xlsx (first sheet) to csv file format
    def xlsx2csv(self, xlsxname, csvname):
        df = pandas.read_excel(xlsxname)
        try:
            df.to_csv(csvname, sep=",", index=False)
        except BaseException as err:
            print(str(err))


    # Write/Append list into xlsx
    def writesheetsfromlist(self, dataframes, xlsxname):
        # Write each dataframe to a different worksheet.
        try:
            book = load_workbook(xlsxname)
            writer = pandas.ExcelWriter(xlsxname, engine='openpyxl')
            writer.book = book
        except:
            writer = pandas.ExcelWriter(xlsxname)

        scnt = 1
        for df in dataframes:
            sheetname = 'Sheet' + str(scnt)
            dframe = pandas.DataFrame(df)
            dframe.to_excel(writer, sheet_name=sheetname)
            scnt += 1

        # Close the Pandas Excel writer and output the Excel file.
        writer.save()

    # Write/Append csv file into xlsx
    def writesheetsfromcsv(self, csvnamelist, xlsxname):
        try:
            book = load_workbook(xlsxname)
            writer = pandas.ExcelWriter(xlsxname, engine='openpyxl')
            writer.book = book
        except:
            writer = pandas.ExcelWriter(xlsxname)

        for csvfile in csvnamelist:
            try:
                df = pandas.read_csv(csvfile)
            except BaseException as err:
                print(err)
            sheetname = csvfile.split('/')[-1]
            dframe = pandas.DataFrame(df)
            dframe.to_excel(writer, sheet_name=sheetname, index=False)

        writer.save()

    # Write/Append query results into xlsx
    def addsheetsfromresultset(self, rows, header, xlsxname, sheetname=None):
        try:
            book = load_workbook(xlsxname)
            writer = pandas.ExcelWriter(xlsxname, engine='openpyxl')
            writer.book = book
        except:
            writer = pandas.ExcelWriter(xlsxname)

        # New empty list called 'result'. This will be written to a file.
        result = list()
        # The row name is the first entry for each entity in the description tuple.
        column_names = list()
        for col in header:
            column_names.append(col)

        for row in rows:
            result.append(row)

        # Write result to file.
        if sheetname == None:
            sheetname = xlsxname.split('/')[-1]

        df = pandas.DataFrame(result, columns=column_names)
        df.to_excel(writer, sheet_name=sheetname, index=False)

        writer.save()