import re
import openpyxl
import os
import psutil
import win32com.client
import unicodedata
from datetime import datetime, timedelta
import time
from random import randint
from docx import Document
from openpyxl.workbook import Workbook
from openpyxl.utils import column_index_from_string

def EmailFormatValidator(email):
    """
    Validates if an email address (parameter: email) is in the correct email format and returns a
       PASS (email in correct format) or FAIL (email not in correct format)
	   
	Robot Framework Usage Example:
			${PASS_FAIL}=	Email Format Validator	  sieqqc@gmail.com	
    """
    if re.match("[\.\w]{2,}[@]\w+[.]\w+",email,re.IGNORECASE):
    	 return "PASS"
    else:
         return "FAIL"

def ContainsOnlyDigits(str_value):
    """
     Validates if a string value (parameter: str_value) contains only digits
       Returns PASS (str_value contains only digits) or FAIL (str_value does not
       contain only digits)
	   
	Robot Framework Usage Example:
			${PASS_FAIL}=	Contains Only Digits	5920782573
    """
    # Using regex() 
    if re.match('^[0-9]*$', str_value): 
       return   'PASS' 
    else: 
       return   'FAIL'

def ConvertStringToList(string): 
    """
     Converts a string (parameter: string) of values INto an array to be used in a list variable
	   Note:  each value in the string parameter must be separated by a space, example:  A B C D
	   
	 Robot Framework Usage Example:
			@{LIST_VARIABLE}=	Convert String To List	A B C D
    """
    ConvLst = list(string.split(" ")) 
    return ConvLst 

def GetValueInString(value, str): 
    """
     Returns all occurrences of a value (parameter: value) contained within a string (parameter: str) or 
	  returns FAIL if the string value (parameter: value) is not contained within the string (parameter: str).
	  
	  Note:  This function is not case sensitive.  The matched value return is in lower case.
	  
	 Robot Framework Usage Example:
					${return_value}=	Get Value In String		is   This is my string
    """
    value = value.lower()
    str = str.lower()
    Match = re.findall(value, str)
    if Match: 
       return    Match
    else: 
       return    'FAIL'
       
def StringFormatValidator(field_format, field_value): 
    """
     Returns PASS if the field_value (parameter: field_value) matches a specified field Regex format (parameter: field_format)
     Returns FAIL if the field_value (parameter: field_value) does not match a specified field Regex format (parameter: field_format)
   
     Robot Framework Usage Example:
				${PASS_FAIL}=    String Format Validator   ^[0-9]{6}-[0-9]{1}$     848567-0
                 Note: must be a string equal to any 6 digits (from 0 to 9) dash any one digit (from 0 to 9)  :
                       016349-0 ; 999999-9 ; 000000-0                    
    """
    Regex_Match = re.match(field_format, field_value)
    if Regex_Match: 
       return    'PASS'
    else: 
       return    'FAIL'

def GetStringPosition(str,fnd): 
    """
    Returns the index (actual position - 1) of the position of the first occurrence of a string (parameter: fnd) contained in the string value passed in (parameter: str)
	
	Note:  This function is not case sensitive
	
	Robot Framework Usage Example:
			${return_value}=	Get String Position	  This is my string		my
    """
    str = str.lower()
    fnd = fnd.lower()
    pos = 0
    ind = str.find(fnd)
    pos += str.find(fnd)
    return pos

def GetUniqueItems(list_of_values):
  """
  Returns the unique list of values from parameter list_of_values (i.e. list variable)
   Note:  This function is case sensitive
  
  Robot Framework Usage Example:
		 @{LIST_OF_VALUES}=    Convert String To List    two three one five one two
         @{UNIQUE_ITEMS}=	   Get Unique Items	         ${LIST_OF_VALUES}
  """
  for x in list_of_values:
      if list_of_values.count(x) > 1:
         list_of_values.remove(x)
  return list_of_values

def CountOccurrenceOfValueInString(string,sub_string):
    """
     Returns the count of the number of times a value (parameter: sub_string) appears in a string (parameter: string)
	 Note:  This function is not case sensitive
	 
	 Robot Framework Usage Example:
				${count_occurence}=	Count Occurrence Of Value In String		One TWO three one five Two	two
    """
    string = string.lower()
    sub_string = sub_string.lower()
    l=len(sub_string)
    count=0
    for i in range(len(string)-len(sub_string)+1):
        if(string[i:i+len(sub_string)] == sub_string ):      
            count+=1
    return count     

def  KillProcess(process_name):
     """
      Kill the process name (parameter: process_name) passed in
	  
	  Robot Framework Usage Example:
	          Kill Process	  chromedriver.exe
	  
     """
     # Iterate over the all the running process
     for proc in psutil.process_iter():
         try:
             # Check if process name contains the given name string (process_name).
             if process_name.lower() in proc.name().lower():
                 proc.kill()
         except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
             pass
            
def RemoveSpecialCharacters(str):
     """
     Removes special characters, including spaces, from a string (parameter: str)
	   and returns the string
	 
	 Robot Framework Usage Example:
          ${special_char_removed}=	Remove Special Characters	${sring_variable}
     """
     alphanumeric = ""
     for character in str:
        if character.isalnum():
           alphanumeric += character
     return	alphanumeric 


def CreateNewWorkbook(filename, sheetname, headings):
	"""
    Creates a workbook object that will be saved at the filename path (parameter: filename) passed in
	  Update the default worksheet name "Sheet" to the value passed in to parameter: sheetname
	  Add the headings to row 1 that are passed in to parameter: headings
	
	Robot Framework Usage Example:
			Create New Workbook   ${workbook_filename_path}  ${sheetname}  ${headings_list}
    """	
	wb = Workbook()   # Create workbook
	
	#Change default worksheet name to the value contained in parameter: sheetname
	ws=wb.get_sheet_by_name('Sheet')
	ws.title = sheetname
	
	# save the workbook
	wb.save(filename)	
	
    # Add the headings to the worksheet	
	wb = openpyxl.load_workbook(filename)
	ws =  wb.get_sheet_by_name(sheetname)
 
	heading_array =  headings.split(";")
	num_headings = len(heading_array)
	
	for x in range(0, num_headings):
		ws.cell(row=1, column=x+1).value = heading_array[x]
	
    # save the workbook	
	wb.save(filename)
    
    
def OpenWorkbook(filename):
    """
    Opens an excel workbook (parameter: filename, which includes the filename path)
	
	Robot Framework Usage Example:
	      Open Workbook		${workbook_filename_path}
		  
		  Note:  ${workbook_filename_path} variable value example = ${EXECDIR}\\Data\\Project_Data_File.xlsx
    """
    wb = openpyxl.load_workbook(filename)
    for sheet in wb:
        print(sheet.title)

def GetDataRowCount(filename, sheetname) :
    """
    Returns the number of rows in a particular worksheet name (parameter: sheetname) of an 
	excel file (parameter: filename, which includes the filename path)
    
    Robot Framework Usage Example:
			${row_count}=	Get Data Row Count  	${workbook_filename_path}  ${worksheet_name}  
          
			Note:  ${workbook_filename_path} variable value example = ${EXECDIR}\\Data\\Project_Data_File.xlsx  
    """
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook.get_sheet_by_name(sheetname)
    row_count = worksheet.max_row-1
    return row_count

def GetDataByRowIndex(excel_row, filename, sheetname) :
    """
    Returns a row of data (into a list variable) from an excel file (parameter: filename) worksheet (parameter: sheetname) for the excel row index 
	  (parameter:  excel_row, which is the excel worksheet row number)
	
	Robot Framework Usage Example:
			 @{DATA_ROW}=	Get Data By Row Index	${excel_row_index_variable}	 ${workbook_filename_path}  ${worksheet_name}  
			 
			 Note:  ${workbook_filename_path} variable value example = ${EXECDIR}\\Data\\Project_Data_File.xlsx  
    """
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook.get_sheet_by_name(sheetname)
    data_row = []                
    excel_row = int(excel_row)   
    for row in worksheet.iter_rows(min_row=excel_row, max_row=excel_row):
        for cell in row:
           #Append column values to the data row list
           data_row.append(cell.value)                                           	     
    return data_row  # return the row of test data
   
def GetNextAvailableDataRow(filename, sheetname, used_col_letter):
     """
     Returns the next available row of data (into a list variable) that is not marked as 'Used' in the column letter (example: column A,B,C,etc.) 
        that is passed in from the excel file (parameter:  filename) worksheet (parameter: sheetname)
		
	 Robot Framework Usage Example:	
             @{DATA_ROW}=	Get Next Available Data Row	${workbook_filename_path}  ${worksheet_name}  ${available_col_letter}

             Note:  ${workbook_filename_path} variable value example = ${EXECDIR}\\Data\\Project_Data_File.xlsx 			 
				
     """
     wb = openpyxl.load_workbook(filename)
     ws =  wb.get_sheet_by_name(sheetname)
     data_row = []                
     excel_col_number = column_index_from_string(used_col_letter)  
     i = 1
     for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
         i = i + 1 
         if ws.cell(i, excel_col_number).value != "Used": 
           available_row = i
           break # exit for loop
     for row in ws.iter_rows(min_row=available_row, max_row=available_row):               
         for cell in row:
           #Append column values to the data row list
           data_row.append(cell.value)  
     ws.cell(row=available_row, column=excel_col_number).value = "Used" # Update 'Available Row' column cell value to 'Used'
     wb.save(filename) # Save the workbook                                                	     
     return data_row  # return the row of test data

def GetAllDataFromExcelSheet(fileName, sheetname) :
    """
    Returns all of the rows of data (into a list variable) from a particular excel file sheetname
	
	Robot Framework Usage Example:
	         @{WORKSHEET_DATA}=		Get All Data From Excel Sheet	${workbook_filename_path}  ${worksheet_name}  

             Note:  ${workbook_filename_path} variable value example = ${EXECDIR}\\Data\\Project_Data_File.xlsx 	
    """
    workbook = openpyxl.load_workbook(fileName)
    worksheet = workbook.get_sheet_by_name(sheetname)
    rowEndIndex = worksheet.max_row
    rowStartIndex = 2  # Start on worksheet row 2, excludes headings row
    data_row = []
    for row in worksheet.iter_rows(min_row=rowStartIndex, max_row=rowEndIndex):
        for cell in row:
            # Append column values to the data row list
            data_row.append(cell.value)
    return data_row

def WriteToExcelFile(filename, sheetname, data_value, row_index, col_index):

     """
     Write a value into a cell in the excel file (parameter: filename) worksheet (parameter: sheetname) 
	   row number (parameter: row_index) and column number (parameter: col_index)
       
	 Robot Framework Usage Example:
	         Write To Excel File	${workbook_filename_path}  ${worksheet_name}  ${data_value}  ${row_index}  ${col_index}  

             Note:  ${workbook_filename_path} variable value example = ${EXECDIR}\\Data\\Project_Data_File.xlsx 	
     """

     wb = openpyxl.load_workbook(filename)
     ws =  wb.get_sheet_by_name(sheetname)

     r = int(row_index)
     c = int(col_index)

     ws.cell(row=r, column=c).value = ''
     ws.cell(row=r, column=c).value = data_value    #enter the data_value in cell row=r and column=c

     wb.save(filename) # Save the workbook

def ReplaceAccents(text):

    """
    Replaces French accents in a string (parameter: text) with the same characters but without the accents
      and returns the string	
	
	Robot Framework Usage Example:
			${string_without_accents}=		Replace Accents	  ${text}
    """

    try:
        text = unicode(text, 'utf-8')
    except NameError: # unicode is a default on python 3
        pass
    text = unicodedata.normalize('NFD', text)\
           .encode('ascii', 'ignore')\
           .decode("utf-8")
    return str(text)


def ReadWordFile(word_file_path): 

    """
    Reads the first sentence of text from a MS Word file (parameter: word_file_path) and returns the MS Word file text 
	
	Robot Framework Usage Example:
			${word_file_text}=	Read Word File	${word_file_path}
			
			Note:  ${word_file_path} variable value example = ${EXECDIR}\\Data\\word_file_name.docx
			
    """  
    document = Document(word_file_path)
    for para in document.paragraphs:
       return (para.text)


def RandomNumber(num_digits):

    """
    Generates a random number having the number of digits inputted in parameter "num_digits" 
	
	Robot Framework Usage Example:
			 ${random_number}=		Random Number	 ${number_digits}
    """ 
    n = int(num_digits)

    range_start = 10**(n-1)

    range_end = (10**n)-1

    return randint(range_start, range_end)

 

def RunRFTestSuite(project_dir, ts_name, ts_subfolders, browser_var, lang_var, run_type_var, tc_row_index_var):

    """
    Runs a Robot Framework Test Suite and moves the test result files to the Reports
    folder relevant to the project main folder (parameter: project_dir) for a given
    test suite (parameters: ts_name), test suite subfolders (parameter: ts_subfolders), and variables
    (parameters: browser_var, lang_var, run_type_var, tc_row_index_var)
    
    Note:  Based on the Robot Framework Project Folder template structure
	
	Robot Framework Usage Example:
	         Run RF Test Suite    ${EXECDIR}  TS_01_US01_Register_Non-CDN_Organization  Test_Suites\\TS_01_P@I_Register_Your_Organization  chrome  en  ${test_case_row_index}
        

    """

    current_date_time = datetime.now()
    timestamp = current_date_time.strftime("%Y%m%d")
    change_dir = "cd " + project_dir

    a = "robot -d "
    b = "Reports" + "\\" + "\\" + ts_name + "-" + timestamp + " --timestampoutputs -r "
    c = "_reports.html -o "
    d = "_output.xml -l "
    e = "_log.html "
    f = "--variable browser:" + browser_var + " --variable lang:" + lang_var + " --variable run_type:" + run_type_var + " --variable test_case_row_index:" + tc_row_index_var + " "
    g = ts_subfolders + "\\" + ts_name + ".robot"

    cmds = a + b + ts_name + c + ts_name + d + ts_name + e + f + g

    print(cmds)

    os.system(change_dir)

    os.system(cmds)